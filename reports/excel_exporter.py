"""
HashInsight Enterprise - Professional Excel Exporter
专业Excel导出器
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.utils import get_column_letter
import io
import logging
from typing import List, Dict

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel报告导出器"""
    
    # 样式配置
    HEADER_FILL = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    PROFIT_FILL = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    LOSS_FILL = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self):
        """初始化Excel导出器"""
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # 移除默认sheet
    
    def create_summary_sheet(self, summary_data: Dict):
        """创建概览工作表"""
        ws = self.wb.create_sheet("概览")
        
        # 标题
        ws['A1'] = "HashInsight 挖矿收益分析报告"
        ws['A1'].font = Font(size=18, bold=True)
        ws.merge_cells('A1:D1')
        
        # 摘要数据
        row = 3
        for key, value in summary_data.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def create_detail_sheet(self, headers: List[str], data: List[List]):
        """创建详细数据工作表"""
        ws = self.wb.create_sheet("详细数据")
        
        # 表头
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        
        # 数据行
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                
                # 条件格式：利润列
                if col_idx == 5 and isinstance(value, (int, float)):
                    if value > 0:
                        cell.fill = self.PROFIT_FILL
                        cell.font = Font(color="FFFFFF")
                    else:
                        cell.fill = self.LOSS_FILL
                        cell.font = Font(color="FFFFFF")
        
        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def create_chart_sheet(self, chart_data: Dict):
        """创建图表工作表"""
        ws = self.wb.create_sheet("图表分析")
        
        # 示例数据（实际应该从chart_data提取）
        ws['A1'] = "日期"
        ws['B1'] = "每日收益"
        
        # 添加示例数据
        for i in range(2, 32):
            ws[f'A{i}'] = f"Day {i-1}"
            ws[f'B{i}'] = chart_data.get('daily_profits', [])[i-2] if len(chart_data.get('daily_profits', [])) >= i-1 else 0
        
        # 创建折线图
        chart = LineChart()
        chart.title = "每日收益趋势"
        chart.y_axis.title = "收益 (USD)"
        chart.x_axis.title = "日期"
        
        data = Reference(ws, min_col=2, min_row=1, max_row=31)
        cats = Reference(ws, min_col=1, min_row=2, max_row=31)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "D2")
    
    def export(self, filename: str = None) -> bytes:
        """导出Excel"""
        if filename:
            self.wb.save(filename)
            return None
        else:
            output = io.BytesIO()
            self.wb.save(output)
            output.seek(0)
            return output.getvalue()
