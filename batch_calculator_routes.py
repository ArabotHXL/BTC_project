"""
Batch calculator routes for handling multiple miner calculations.
"""
from flask import Blueprint, request, jsonify, render_template, session, render_template_string, send_file, make_response, Response
from auth import login_required
from decorators import check_miner_limit, get_user_plan, UpgradeRequired, require_feature
from common.rbac import requires_module_access, Module, AccessLevel
from mining_calculator import calculate_mining_profitability, MINER_DATA
from api_client import get_btc_price_with_fallback, get_network_stats_with_fallback
from optimized_batch_processor import batch_processor
from fast_batch_processor import fast_batch_processor
from models import MinerModel
import logging
import io
import os
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter

# Create blueprint
batch_calculator_bp = Blueprint('batch_calculator', __name__)

logger = logging.getLogger(__name__)


@batch_calculator_bp.route('/batch-calculator')
@login_required
def batch_calculator():
    """Display the batch calculator interface."""
    try:
        # Import flask's g and translation function here to ensure proper context
        from flask import g
        from translations import get_translation
        
        user_id = session.get('user_id')
        user_plan_raw = get_user_plan(user_id)
        
        # Use the actual user plan from database
        user_plan = user_plan_raw
        
        # Only create default plan if we don't have a valid plan object
        if not user_plan or not hasattr(user_plan, 'name'):
            class DefaultPlan:
                id = 'free'
                name = 'Free'
                max_miners = 10
                price = 0
                allow_advanced_analytics = False
                allow_api = False
                allow_scenarios = False
            user_plan = DefaultPlan()
        
        # Get current language from g (set by before_request) or session
        current_lang = getattr(g, 'language', session.get('language', 'zh'))
        
        # Ensure language is saved to session
        session['language'] = current_lang
        
        # Set g.language if not already set
        if not hasattr(g, 'language'):
            g.language = current_lang
        
        plan_name = getattr(user_plan, 'name', user_plan) if user_plan else 'Free'
        logger.info(f"Rendering batch calculator with plan: {plan_name}, language: {current_lang}")
        
        # Get miner models from database 
        miner_models_dict = {}
        
        # Load miner models directly from database
        try:
            from models import db
            from sqlalchemy import text
            
            # Handle any failed transaction by rolling back
            try:
                db.session.rollback()
            except:
                pass
            
            # Query all active miner models from database
            query = text("""
                SELECT model_name, reference_hashrate, reference_power, reference_price, manufacturer, reference_efficiency
                FROM miner_models 
                WHERE is_active = true 
                ORDER BY model_name
            """)
            
            result = db.session.execute(query)
            
            for row in result:
                model_name = row[0]
                miner_models_dict[model_name] = {
                    'hashrate': float(row[1]) if row[1] else 0,
                    'power': int(row[2]) if row[2] else 0,
                    'price': float(row[3]) if row[3] else 0,
                    'manufacturer': row[4] if row[4] else '',
                    'efficiency': float(row[5]) if row[5] else 0
                }
            
            # Commit the transaction
            db.session.commit()
            logger.info(f"Successfully loaded {len(miner_models_dict)} miner models from database")
            
        except Exception as e:
            logger.error(f"Failed to load miner models from database: {e}")
            # Fallback to MINER_DATA if database fails
            try:
                from mining_calculator import MINER_DATA
                for model_name, specs in MINER_DATA.items():
                    miner_models_dict[model_name] = {
                        'hashrate': float(specs.get('hashrate', 0)),  # 正确的字段名
                        'power': int(specs.get('power_watt', 0)),     # 正确的字段名
                        'price': float(specs.get('price', 2500)),     # 默认价格
                        'manufacturer': specs.get('manufacturer', 'Unknown'),
                        'efficiency': float(specs.get('efficiency', 30))  # 默认效率
                    }
                logger.info(f"Successfully loaded {len(miner_models_dict)} models from MINER_DATA fallback")
                
                # Debug: print first few entries to verify data
                if miner_models_dict:
                    sample_model = list(miner_models_dict.keys())[0]
                    logger.info(f"Sample model data: {sample_model} = {miner_models_dict[sample_model]}")
                    
                logger.warning(f"Using MINER_DATA fallback: {len(miner_models_dict)} models")
            except:
                # Ultimate fallback
                miner_models_dict = {
                    'Antminer S19 Pro': {'hashrate': 110, 'power': 3250, 'price': 2500, 'manufacturer': 'Bitmain', 'efficiency': 29.5},
                    'Antminer S21': {'hashrate': 200, 'power': 3550, 'price': 3200, 'manufacturer': 'Bitmain', 'efficiency': 17.8},
                    'WhatsMiner M53S': {'hashrate': 226, 'power': 6554, 'price': 4500, 'manufacturer': 'MicroBT', 'efficiency': 29.0}
                }
                logger.warning(f"Using minimal fallback: {len(miner_models_dict)} models")
        
        # Pass session data to template with translation function and miner models
        template_data = {
            'user_plan': user_plan,
            'current_lang': current_lang,
            'session': session,
            'miner_models': miner_models_dict,
            't': lambda text: get_translation(text, to_lang=current_lang or 'zh')
        }
        
        return render_template('batch_calculator.html', **template_data)
    
    except Exception as e:
        logger.error(f"Error loading batch calculator: {e}", exc_info=True)
        # Return a working basic page
        return render_template_string('''
<!DOCTYPE html>
<html lang="zh-CN" data-bs-theme="dark">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>批量挖矿计算器</title>
    <link rel="stylesheet" href="https://cdn.replit.com/agent/bootstrap-agent-dark-theme.min.css">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.1/font/bootstrap-icons.css">
</head>
<body>
    <nav class="navbar navbar-expand-lg navbar-dark bg-dark border-bottom">
        <div class="container">
            <a class="navbar-brand fw-bold" href="/">
                <i class="bi bi-calculator me-2"></i>BTC挖矿计算器
            </a>
            <div class="d-flex gap-2">
                <a href="/" class="btn btn-outline-light btn-sm">
                    <i class="bi bi-arrow-left me-1"></i>返回计算器
                </a>
            </div>
        </div>
    </nav>

    <div class="container mt-4">
        <div class="row mb-4">
            <div class="col-12">
                <h2 class="mb-1">批量挖矿计算器</h2>
                <p class="text-muted mb-0">一次性计算多台矿机的收益率</p>
                
                <div class="alert alert-info mt-3">
                    <i class="bi bi-info-circle me-2"></i>
                    <strong>系统维护中</strong>
                    <p class="mb-0">批量计算器正在升级优化，请稍后再试。您可以先使用主页的单台矿机计算功能。</p>
                </div>
                
                <div class="text-center mt-4">
                    <a href="/" class="btn btn-primary">返回主计算器</a>
                    <a href="/billing/plans" class="btn btn-outline-secondary ms-2">查看订阅计划</a>
                </div>
            </div>
        </div>
    </div>
</body>
</html>
        ''')


@batch_calculator_bp.route('/api/batch-calculate', methods=['POST'])
@login_required
@requires_module_access(Module.ANALYTICS_BATCH_CALC)
def batch_calculate():
    """Process batch mining calculations with optimized performance for large datasets."""
    try:
        data = request.get_json()
        
        if not data or 'miners' not in data:
            return jsonify({
                'success': False,
                'error': 'invalid_request',
                'message': 'Invalid request data'
            }), 400
        
        miners = data['miners']
        settings = data.get('settings', {})
        
        # Calculate total miner count
        total_miners = sum(miner.get('quantity', 1) for miner in miners)
        
        # Skip the ultra-fast processor call since it's not implemented correctly
        # Use the standard batch processing logic instead
        logger.info(f"Processing batch calculation: {total_miners} miners")
        
        # Fallback to optimized processor for large datasets
        if total_miners > 1000:
            logger.info(f"Processing large batch: {total_miners} miners - Using optimized processor")
            
            # 使用优化处理器处理大批量数据
            result = batch_processor.process_large_batch(miners, use_real_time_data=True)
            if result['success']:
                logger.info(f"优化处理器成功处理: {result['summary']['total_miners']} 矿机")
                return jsonify(result)
            else:
                logger.error(f"优化处理器失败: {result.get('error', 'Unknown error')}")
                # 如果优化处理器失败，继续使用标准处理
            
        # Optimized batch processing
        results = []
        total_daily_profit = 0
        total_daily_revenue = 0
        total_daily_cost = 0
        batch_size = 100  # Process in chunks to manage memory
        
        # Group identical miners to optimize calculations (including hashrate, decay_rate, and machine_price for ROI)
        # 但保留矿机编号信息以便在结果中显示
        miner_groups = {}
        miner_numbers_map = {}  # 记录每组的矿机编号
        miner_price_map = {}     # 记录每组的矿机价格
        
        for miner in miners:
            key = (
                miner.get('model', 'Antminer S19 Pro'),
                float(miner.get('power_consumption', 3250)),
                float(miner.get('electricity_cost', 0.08)),
                float(miner.get('hashrate', 0)),
                float(miner.get('decay_rate', 0)),
                float(miner.get('machine_price', 0))  # 添加机器价格到分组key中
            )
            miner_number = miner.get('miner_number', '')
            machine_price = float(miner.get('machine_price', 0))
            
            if key not in miner_groups:
                miner_groups[key] = 0
                miner_numbers_map[key] = []
                miner_price_map[key] = machine_price
            
            miner_groups[key] += int(miner.get('quantity', 1))
            if miner_number:
                miner_numbers_map[key].append(miner_number)
        
        logger.info(f"Optimized {len(miners)} entries into {len(miner_groups)} unique groups")
        
        # 预获取共享数据以加速批量计算
        start_time = time.time()
        try:
            # 获取共享的网络数据 - 使用更快的数据库缓存
            from db import get_db_connection
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # 从market_analytics表获取最新数据（更快）
            cursor.execute("""
                SELECT btc_price, network_difficulty, network_hashrate, block_reward 
                FROM market_analytics 
                ORDER BY recorded_at DESC 
                LIMIT 1
            """)
            
            result = cursor.fetchone()
            if result:
                shared_network_data = {
                    'btc_price': float(result[0]),
                    'network_difficulty': float(result[1]),
                    'network_hashrate': float(result[2]),
                    'block_reward': float(result[3])
                }
                data_fetch_time = time.time() - start_time
                logger.info(f"Fast database data fetched in {data_fetch_time:.3f}s: BTC=${result[0]}, Hashrate={result[2]}EH/s")
            else:
                # fallback to API if database is empty
                btc_price = get_btc_price_with_fallback()
                network_stats = get_network_stats_with_fallback()
                shared_network_data = {
                    'btc_price': btc_price,
                    'network_difficulty': network_stats.get('difficulty'),
                    'network_hashrate': network_stats.get('hashrate'),
                    'block_reward': network_stats.get('block_reward', 3.125)
                }
                data_fetch_time = time.time() - start_time
                logger.info(f"Fallback API data fetched in {data_fetch_time:.3f}s")
                
        except Exception as e:
            logger.warning(f"Failed to fetch shared data: {e}, using individual API calls")
            shared_network_data = None
        
        # Calculate once per unique group
        for groupKey, quantity in miner_groups.items():
            (model, power_consumption, electricity_cost, hashrate, decay_rate, machine_price) = groupKey
            total_investment = machine_price * quantity  # 计算总投资成本
            
            try:
                # Single calculation for the entire group with batch optimization
                # 如果有自定义算力，直接使用算力值；否则使用矿机型号
                if hashrate > 0:
                    # 使用自定义算力，传递预获取的共享数据
                    calc_result = calculate_mining_profitability(
                        hashrate=hashrate * quantity,  # 总算力
                        power_consumption=power_consumption * quantity,  # 总功耗
                        electricity_cost=electricity_cost,
                        host_investment=total_investment,  # 正确的参数名：host_investment
                        btc_price=shared_network_data.get('btc_price') if shared_network_data else None,
                        difficulty=shared_network_data.get('network_difficulty') if shared_network_data else None,
                        block_reward=shared_network_data.get('block_reward') if shared_network_data else None,
                        manual_network_hashrate=shared_network_data.get('network_hashrate') if shared_network_data else None,
                        use_real_time_data=False  # 禁用实时API调用
                    )
                else:
                    # 使用矿机型号默认值，传递预获取的共享数据
                    calc_result = calculate_mining_profitability(
                        power_consumption=power_consumption,
                        electricity_cost=electricity_cost,
                        miner_model=model,
                        miner_count=quantity,
                        host_investment=total_investment,  # 正确的参数名：host_investment
                        btc_price=shared_network_data.get('btc_price') if shared_network_data else None,
                        difficulty=shared_network_data.get('network_difficulty') if shared_network_data else None,
                        block_reward=shared_network_data.get('block_reward') if shared_network_data else None,
                        manual_network_hashrate=shared_network_data.get('network_hashrate') if shared_network_data else None,
                        use_real_time_data=False  # 禁用实时API调用
                    )
                
                # 获取这组矿机的编号
                group_numbers = miner_numbers_map.get(groupKey, [])
                display_number = ', '.join(group_numbers) if group_numbers else '-'
                
                # 从ROI数据中提取回收期并转换为天数
                roi_data = calc_result.get('roi', {}).get('host', {})
                payback_months = roi_data.get('payback_period_months', 0)
                
                # 调试日志
                logger.info(f"ROI calculation debug - ROI data keys: {list(roi_data.keys()) if roi_data else 'No ROI data'}")
                logger.info(f"ROI calculation debug - Payback months: {payback_months}")
                
                # 如果payback_months为None，使用简单的投资/月利润计算
                if not payback_months and total_investment > 0:
                    # 尝试从不同的数据结构中获取月利润
                    monthly_profit = (calc_result.get('monthly_profit', 0) or 
                                    calc_result.get('profit', {}).get('monthly', 0))
                    if monthly_profit > 0:
                        payback_months = total_investment / monthly_profit
                        logger.info(f"ROI fallback calculation - Investment: ${total_investment}, Monthly profit: ${monthly_profit}, Calculated months: {payback_months}")
                
                roi_days = int(payback_months * 30.44) if payback_months and payback_months > 0 else 0  # 平均每月30.44天
                
                result_entry = {
                    'miner_number': display_number,  # 添加矿机编号
                    'model': model,
                    'quantity': quantity,
                    'power_consumption': power_consumption,
                    'electricity_cost': electricity_cost,
                    'machine_price': machine_price,  # 添加单台矿机价格
                    'total_machine_cost': total_investment,  # 添加总投资成本
                    'daily_profit': calc_result.get('daily_profit', 0),
                    'daily_revenue': calc_result.get('daily_revenue', 0),
                    'daily_cost': calc_result.get('daily_cost', 0),
                    'monthly_profit': (calc_result.get('monthly_profit', 0) or 
                                     calc_result.get('profit', {}).get('monthly', 0)),
                    'roi_days': roi_days,  # 正确计算的ROI天数
                    'hash_rate': calc_result.get('hash_rate', hashrate),
                    'decay_rate': decay_rate
                }
                
                results.append(result_entry)
                
                # Add to totals
                total_daily_profit += result_entry['daily_profit']
                total_daily_revenue += result_entry['daily_revenue']
                total_daily_cost += result_entry['daily_cost']
                
            except Exception as e:
                logger.error(f"Error calculating for miner group {model}: {e}")
                continue
        
        # Create summary with reduced precision to save memory
        summary = {
            'total_miners': total_miners,
            'total_daily_profit': round(total_daily_profit, 2),
            'total_daily_revenue': round(total_daily_revenue, 2),
            'total_daily_cost': round(total_daily_cost, 2),
            'total_monthly_profit': round(total_daily_profit * 30, 2),
            'unique_groups': len(results),
            'average_roi_days': round(sum(r.get('roi_days', 0) for r in results) / len(results), 1) if results else 0
        }
        
        logger.info(f"Optimized batch calculation completed: {len(results)} unique groups, {total_miners} total miners")
        
        return jsonify({
            'success': True,
            'results': results,
            'summary': summary,
            'optimization_info': {
                'original_entries': len(miners),
                'optimized_groups': len(results),
                'total_miners': total_miners,
                'memory_optimized': total_miners > 1000
            }
        })
        
    except UpgradeRequired as e:
        return jsonify({
            'success': False,
            'error': 'upgrade_required',
            'message': str(e),
            'current_plan': 'Free'
        }), 402
        
    except Exception as e:
        logger.error(f"Batch calculation error: {e}")
        return jsonify({
            'success': False,
            'error': 'calculation_failed',
            'message': 'Batch calculation failed. Please try again.'
        }), 500


@batch_calculator_bp.route('/api/export/batch-csv', methods=['POST'])
@login_required
@requires_module_access(Module.REPORT_EXCEL)
def export_batch_csv():
    """Export batch calculation results to CSV.
    
    RBAC权限 (REPORT_EXCEL):
    - All except Guest: FULL
    - Guest: NONE
    """
    try:
        data = request.get_json()
        results = data.get('results', [])
        
        if not results:
            return jsonify({
                'success': False,
                'message': 'No results to export'
            }), 400
        
        # Generate CSV content with 3 decimal places precision
        csv_lines = ['Model,Quantity,Daily Revenue,Daily Cost,Daily Profit,Monthly Profit,ROI Days,Hash Rate (TH/s),Power (W),Machine Price,Total Cost,Daily BTC,Monthly BTC']
        
        for result in results:
            line = ','.join([
                f'"{result.get("model", "")}"',
                str(result.get('quantity', 0)),
                f"{result.get('daily_revenue', 0):.2f}",
                f"{result.get('daily_cost', 0):.2f}",
                f"{result.get('daily_profit', 0):.2f}",
                f"{result.get('monthly_profit', 0):.2f}",
                str(result.get('roi_days', 0)),
                f"{result.get('hash_rate', 0):.0f}",
                f"{result.get('power_consumption', 0):.0f}",
                f"{result.get('machine_price', 0):.2f}",
                f"{result.get('total_machine_cost', 0):.2f}",
                f"{result.get('daily_btc', 0):.8f}",
                f"{result.get('monthly_btc', 0):.8f}"
            ])
            csv_lines.append(line)
        
        csv_content = '\n'.join(csv_lines)
        
        # Return CSV file as download
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename="batch_mining_results_{len(results)}_miners.csv"'
        return response
        
    except Exception as e:
        logger.error(f"CSV export error: {e}")
        return jsonify({
            'success': False,
            'message': 'Export failed'
        }), 500


@batch_calculator_bp.route('/api/export/batch-excel', methods=['POST'])
@login_required
@requires_module_access(Module.REPORT_EXCEL)
def export_batch_excel():
    """Export batch calculation results to Excel (Basic+ plans only).
    
    RBAC权限 (REPORT_EXCEL):
    - All except Guest: FULL
    - Guest: NONE
    """
    try:
        user_id = session.get('user_id')
        plan = get_user_plan(user_id)
        
        # Check if user has Excel export permissions
        if not plan or plan == 'free':
            return jsonify({
                'success': False,
                'error': 'upgrade_required',
                'message': 'Excel export requires Basic plan or higher'
            }), 402
        
        data = request.get_json()
        results = data.get('results', [])
        
        if not results:
            return jsonify({
                'success': False,
                'message': 'No results to export'
            }), 400
        
        # Generate Excel file using openpyxl
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        if worksheet is not None:
            worksheet.title = "Mining Results"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers with 3 decimal places precision
        headers = [
            'Miner Model', 'Quantity', 'Daily Revenue (USD)', 'Daily Cost (USD)', 
            'Daily Profit (USD)', 'Monthly Profit (USD)', 'Annual ROI (%)', 
            'Payback Days', 'Hash Rate (TH/s)', 'Power (W)', 
            'Daily BTC', 'Monthly BTC'
        ]
        
        # Set headers
        if worksheet is not None:
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
        
        # Add data rows with 3 decimal places
        if worksheet is not None:
            for row, result in enumerate(results, 2):
                data_row = [
                    result.get('model', ''),
                    result.get('quantity', 0),
                    round(result.get('daily_revenue', 0), 3),
                    round(result.get('daily_cost', 0), 3),
                    round(result.get('daily_profit', 0), 3),
                    round(result.get('monthly_profit', 0), 3),
                    round(result.get('annual_roi', 0), 3),
                    round(result.get('roi_days', 0), 1),      # ROI天数1位小数
                    round(result.get('hash_rate', 0), 1),   # 算力1位小数
                    round(result.get('power_consumption', 0), 0),  # 功耗整数
                    round(result.get('daily_btc', 0), 8),
                    round(result.get('monthly_btc', 0), 8)
                ]
                
                for col, value in enumerate(data_row, 1):
                    cell = worksheet.cell(row=row, column=col, value=value)
                    cell.border = border
                    # 设置不同列的数字格式
                    if isinstance(value, (int, float)) and col > 2:
                        if col in [11, 12]:  # Daily BTC, Monthly BTC
                            cell.number_format = '0.00000000'  # 8位小数
                        elif col in [9, 10]:  # Hash Rate, Power
                            cell.number_format = '0.0'  # 1位小数
                        elif col == 8:  # Payback Days
                            cell.number_format = '0.0'  # 1位小数
                        else:  # 其他金额字段
                            cell.number_format = '0.000'  # 3位小数
        
        # Auto-adjust column widths
        if worksheet is not None:
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"batch_mining_results_{len(results)}_miners_{timestamp}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return jsonify({
            'success': False,
            'message': 'Export failed'
        }), 500


@batch_calculator_bp.route('/api/export/batch-pdf', methods=['POST'])
@login_required
@requires_module_access(Module.REPORT_PDF)
def export_batch_pdf():
    """Export batch calculation results to PDF (Pro+ plans only).
    
    RBAC权限 (REPORT_PDF):
    - All except Guest: FULL
    - Guest: READ (demo only)
    """
    try:
        user_id = session.get('user_id')
        plan = get_user_plan(user_id)
        
        # Check if user has PDF export permissions  
        if not plan or plan == 'free':
            return jsonify({
                'success': False,
                'error': 'upgrade_required',
                'message': 'PDF export requires Pro plan or higher'
            }), 402
        
        data = request.get_json()
        results = data.get('results', [])
        summary = data.get('summary', {})
        
        if not results:
            return jsonify({
                'success': False,
                'message': 'No results to export'
            }), 400
        
        # Generate PDF using reportlab
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        import datetime
        
        # Create buffer for PDF
        buffer = BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            textColor=colors.HexColor('#1f2937'),
            alignment=1  # Center
        )
        
        story.append(Paragraph("Bitcoin Mining Profitability Report", title_style))
        story.append(Paragraph(f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Summary
        summary = data.get('summary', {})
        summary_data = [
            ['Summary', ''],
            ['Total Miners', f"{summary.get('total_miners', 0):,}"],
            ['Daily Profit', f"${summary.get('total_daily_profit', 0):,.2f}"],
            ['Monthly Profit', f"${summary.get('total_monthly_profit', 0):,.2f}"],
            ['Average ROI Days', f"{summary.get('average_roi_days', 'N/A')} days" if summary.get('average_roi_days', 999999) < 999999 else 'N/A']
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Results table
        story.append(Paragraph("Individual Results", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        # Table headers
        table_data = [['#', 'Model', 'Qty', 'Daily Profit', 'ROI Days', 'Hashrate (TH/s)']]
        
        # Sort results by daily profit (descending)
        sorted_results = sorted(results, key=lambda x: x.get('daily_profit', 0), reverse=True)
        
        for result in sorted_results[:50]:  # Limit to first 50 for PDF size
            table_data.append([
                result.get('miner_number', '-'),
                result.get('model', '-'),
                str(result.get('quantity', 0)),
                f"${result.get('daily_profit', 0):.2f}",
                f"{result.get('roi_days', 'N/A')} days" if result.get('roi_days', 999999) < 999999 else 'N/A',
                f"{result.get('hash_rate', 0):.1f}"
            ])
        
        # Create table
        results_table = Table(table_data, colWidths=[0.8*inch, 1.5*inch, 0.7*inch, 1.2*inch, 1.2*inch, 1.1*inch])
        results_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(results_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        # Generate filename and return PDF directly as response
        timestamp = int(datetime.datetime.now().timestamp() * 1000)
        filename = f"mining_batch_results_{timestamp}.pdf"
        
        # Create response with PDF content
        response = Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'application/pdf'
            }
        )
        
        return response
        
    except Exception as e:
        logger.error(f"PDF export error: {e}")
        return jsonify({
            'success': False,
            'message': 'Export failed'
        }), 500


@batch_calculator_bp.route('/api/batch-calculate-optimized', methods=['POST'])
@login_required
@requires_module_access(Module.ANALYTICS_BATCH_CALC)
def batch_calculate_optimized():
    """优化的批量计算接口，专门处理大量数据 (5000+ miners)"""
    try:
        data = request.get_json()
        
        if not data or 'miners' not in data:
            return jsonify({
                'success': False,
                'error': 'invalid_request',
                'message': 'Invalid request data'
            }), 400
        
        miners = data['miners']
        settings = data.get('settings', {})
        
        # 计算总矿机数量
        total_miners = sum(miner.get('quantity', 1) for miner in miners)
        
        logger.info(f"优化批量计算: {len(miners)} 条目, {total_miners} 矿机")
        
        # 检查权限
        allowed = check_miner_limit(total_miners)
        if not allowed:
            return jsonify({
                'success': False,
                'error': 'upgrade_required',
                'message': f'您的计划不支持 {total_miners} 台矿机。Pro 计划支持无限制批量计算。'
            }), 402
        
        # 使用优化处理器
        result = batch_processor.process_large_batch(
            miners, 
            use_real_time_data=settings.get('use_real_time_data', True)
        )
        
        if result['success']:
            logger.info(f"优化计算完成: {result['summary']['total_miners']} 矿机, 日收益 ${result['summary']['total_daily_profit']:,.2f}")
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"优化批量计算错误: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'calculation_failed',
            'message': '优化批量计算失败，请重试'
        }), 500


# 新增：用户矿机自动保存/加载功能
@batch_calculator_bp.route('/api/user-miners', methods=['GET'])
@login_required
def get_user_miners():
    """获取用户保存的矿机配置"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({
                'success': False,
                'error': 'authentication_required',
                'message': 'Please login first'
            }), 401

        from models import UserMiner, db
        
        # 获取用户的活跃矿机配置
        user_miners = UserMiner.query.filter_by(
            user_id=user_id,
            status='active'
        ).order_by(UserMiner.created_at.desc()).all()
        
        # 转换为前端需要的格式
        miners_data = []
        for miner in user_miners:
            miners_data.append({
                'id': miner.id,
                'model': miner.miner_model.model_name if miner.miner_model else '',
                'quantity': miner.quantity,
                'power': miner.actual_power,
                'price': miner.actual_price,
                'electricity': miner.electricity_cost,
                'hashrate': miner.actual_hashrate,
                'decayRate': miner.decay_rate_monthly,
                'custom_name': miner.custom_name,
                'location': miner.location
            })
        
        return jsonify({
            'success': True,
            'miners': miners_data,
            'count': len(miners_data)
        })
        
    except Exception as e:
        logger.error(f"Failed to load user miners: {e}")
        return jsonify({
            'success': False,
            'error': 'load_failed',
            'message': 'Failed to load saved miners'
        }), 500


@batch_calculator_bp.route('/api/user-miners', methods=['POST'])
@login_required  
def save_user_miners():
    """保存用户矿机配置（自动保存）"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({
                'success': False,
                'error': 'authentication_required',
                'message': 'Please login first'
            }), 401

        data = request.get_json()
        if not data or 'miners' not in data:
            return jsonify({
                'success': False,
                'error': 'invalid_data',
                'message': 'Invalid miner data'
            }), 400

        miners = data['miners']
        from models import UserMiner, MinerModel, db
        
        # 清除用户现有的自动保存配置（保留手动命名的配置）
        UserMiner.query.filter_by(
            user_id=user_id,
            custom_name=None  # 只删除没有自定义名称的自动保存配置
        ).delete()
        
        saved_count = 0
        
        # 保存新的矿机配置
        for miner_data in miners:
            # 查找对应的矿机型号
            miner_model = MinerModel.query.filter_by(
                model_name=miner_data.get('model', ''),
                is_active=True
            ).first()
            
            if not miner_model:
                logger.warning(f"Miner model not found: {miner_data.get('model', '')}")
                continue
            
            # 创建用户矿机记录
            user_miner = UserMiner(
                user_id=user_id,
                miner_model_id=miner_model.id,
                quantity=int(miner_data.get('quantity', 1)),
                actual_hashrate=float(miner_data.get('hashrate', 0)),
                actual_power=int(miner_data.get('power', 0)),
                actual_price=float(miner_data.get('price', 0)),
                electricity_cost=float(miner_data.get('electricity', 0)),
                decay_rate_monthly=float(miner_data.get('decayRate', 0.5)),
                custom_name=None,  # 自动保存的配置不设置自定义名称
                status='active'
            )
            
            db.session.add(user_miner)
            saved_count += 1
        
        db.session.commit()
        
        logger.info(f"Auto-saved {saved_count} miner configurations for user {user_id}")
        
        return jsonify({
            'success': True,
            'message': f'Configuration auto-saved',
            'saved_count': saved_count
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Failed to save user miners: {e}")
        return jsonify({
            'success': False,
            'error': 'save_failed',
            'message': 'Failed to save configuration'
        }), 500


@batch_calculator_bp.route('/api/import-my-miners', methods=['GET'])
@login_required
def import_my_miners():
    """从托管系统导入所有活跃矿机到批量计算器"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({
                'success': False,
                'error': 'authentication_required',
                'message': 'Please login first'
            }), 401

        try:
            user_id_int = int(user_id)
        except (ValueError, TypeError):
            return jsonify({
                'success': False,
                'error': 'invalid_user',
                'message': 'Invalid user ID'
            }), 400

        from models import HostingMiner, HostingSite, MinerModel, db
        from sqlalchemy import func

        hosting_results = db.session.query(
            MinerModel.model_name,
            func.count(HostingMiner.id).label('count'),
            func.avg(HostingMiner.actual_hashrate).label('avg_hashrate'),
            func.avg(HostingMiner.actual_power).label('avg_power'),
            HostingSite.electricity_rate
        ).join(
            MinerModel, HostingMiner.miner_model_id == MinerModel.id
        ).join(
            HostingSite, HostingMiner.site_id == HostingSite.id
        ).filter(
            HostingMiner.customer_id == user_id_int,
            HostingMiner.status.in_(['active', 'online'])
        ).group_by(
            MinerModel.model_name, HostingSite.electricity_rate
        ).all()

        miners_data = []
        for row in hosting_results:
            miners_data.append({
                'model': row.model_name,
                'quantity': row.count,
                'power': round(row.avg_power or 0),
                'price': 0,
                'electricity': row.electricity_rate or 0.08,
                'hashrate': round(row.avg_hashrate or 0, 1),
                'decayRate': 0.5
            })

        if not miners_data:
            return jsonify({
                'success': True,
                'miners': [],
                'count': 0,
                'message': '托管系统中没有找到活跃矿机'
            })

        return jsonify({
            'success': True,
            'miners': miners_data,
            'count': len(miners_data)
        })

    except Exception as e:
        logger.error(f"Failed to import user miners: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'import_failed',
            'message': 'Failed to import miners from your account'
        }), 500