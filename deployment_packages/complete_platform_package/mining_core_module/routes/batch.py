"""
批量计算路由模块
Batch Calculator Routes Module

提供批量挖矿计算API和页面路由，无用户认证依赖
"""
import logging
import io
import os
import time
from datetime import datetime
from flask import Blueprint, request, jsonify, render_template, send_file, make_response
from models import MinerModel, NetworkSnapshot, db
from mining_calculator import calculate_mining_profitability
from utils.optimized_batch_processor import OptimizedBatchProcessor

# 导入Excel相关库
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl not available, Excel export disabled")

# 创建蓝图
batch_bp = Blueprint('batch', __name__)

logger = logging.getLogger(__name__)

# 配置限制（简化版，无用户认证）
MAX_BATCH_SIZE = 1000  # 最大批量计算数量
DEFAULT_BATCH_LIMIT = 100  # 默认批量限制

@batch_bp.route('/')
def index():
    """批量计算器主页面"""
    try:
        # 获取可用的矿机列表
        miners = MinerModel.get_active_miners()
        
        # 获取最新网络数据
        latest_network = NetworkSnapshot.query.order_by(
            NetworkSnapshot.recorded_at.desc()
        ).first()
        
        # 准备矿机数据字典
        miner_models_dict = {}
        for miner in miners:
            miner_models_dict[miner.model_name] = {
                'hashrate': miner.reference_hashrate,
                'power': miner.reference_power,
                'price': miner.reference_price or 2500,
                'manufacturer': miner.manufacturer,
                'efficiency': miner.reference_efficiency or 30
            }
        
        return render_template('batch/index.html', 
                             miners=miner_models_dict,
                             network_data=latest_network,
                             max_batch_size=MAX_BATCH_SIZE)
                             
    except Exception as e:
        logger.error(f"加载批量计算器页面失败: {e}")
        return render_template('errors/500.html'), 500

@batch_bp.route('/api/batch-calculate', methods=['POST'])
def batch_calculate():
    """处理批量挖矿计算"""
    try:
        data = request.get_json()
        
        if not data or 'miners' not in data:
            return jsonify({
                'success': False,
                'error': 'invalid_request',
                'message': '无效的请求数据'
            }), 400
        
        miners = data['miners']
        settings = data.get('settings', {})
        
        # 计算总矿机数量
        total_miners = sum(miner.get('quantity', 1) for miner in miners)
        
        # 检查数量限制
        if total_miners > MAX_BATCH_SIZE:
            return jsonify({
                'success': False,
                'error': 'limit_exceeded',
                'message': f'批量计算数量超出限制。最大允许: {MAX_BATCH_SIZE}台',
                'max_allowed': MAX_BATCH_SIZE,
                'requested': total_miners
            }), 400
        
        logger.info(f"开始处理批量计算: {total_miners} 台矿机")
        
        # 使用优化的批量处理器
        if total_miners > 500:
            processor = OptimizedBatchProcessor()
            result = processor.process_large_batch(miners, use_real_time_data=True)
            
            if result['success']:
                logger.info(f"优化处理器成功处理: {result['summary']['total_miners']} 矿机")
                return jsonify(result)
            else:
                logger.warning(f"优化处理器失败: {result.get('error', 'Unknown error')}")
                # 继续使用标准处理
        
        # 标准批量处理逻辑
        results = []
        total_daily_profit = 0
        total_daily_revenue = 0
        total_daily_cost = 0
        
        # 分组相同配置的矿机以优化计算
        miner_groups = {}
        miner_numbers_map = {}
        
        for miner in miners:
            key = (
                miner.get('model', 'Antminer S19 Pro'),
                float(miner.get('power_consumption', 3250)),
                float(miner.get('electricity_cost', 0.06)),
                float(miner.get('hashrate', 0)),
                float(miner.get('decay_rate', 0)),
                float(miner.get('machine_price', 2500))
            )
            miner_number = miner.get('miner_number', '')
            
            if key not in miner_groups:
                miner_groups[key] = 0
                miner_numbers_map[key] = []
            
            miner_groups[key] += int(miner.get('quantity', 1))
            if miner_number:
                miner_numbers_map[key].append(miner_number)
        
        logger.info(f"优化: {len(miners)} 条记录分组为 {len(miner_groups)} 个计算组")
        
        # 获取共享网络数据
        shared_network_data = get_shared_network_data()
        
        # 对每个分组进行计算
        for group_key, quantity in miner_groups.items():
            (model, power_consumption, electricity_cost, hashrate, decay_rate, machine_price) = group_key
            total_investment = machine_price * quantity
            
            try:
                # 执行挖矿收益计算
                if hashrate > 0:
                    # 使用自定义算力
                    calc_result = calculate_mining_profitability(
                        hashrate=hashrate * quantity,
                        power_consumption=power_consumption * quantity,
                        electricity_cost=electricity_cost,
                        custom_btc_price=shared_network_data.get('btc_price'),
                        custom_difficulty=shared_network_data.get('network_difficulty'),
                        use_real_time_data=False
                    )
                else:
                    # 使用矿机型号默认值
                    calc_result = calculate_mining_profitability(
                        miner_model=model,
                        miner_count=quantity,
                        electricity_cost=electricity_cost,
                        custom_btc_price=shared_network_data.get('btc_price'),
                        custom_difficulty=shared_network_data.get('network_difficulty'),
                        use_real_time_data=False
                    )
                
                # 获取矿机编号
                group_numbers = miner_numbers_map.get(group_key, [])
                display_number = ', '.join(group_numbers) if group_numbers else '-'
                
                # 计算ROI
                monthly_profit = calc_result.get('monthly_profit', 0)
                roi_days = 0
                if monthly_profit > 0 and total_investment > 0:
                    roi_months = total_investment / monthly_profit
                    roi_days = int(roi_months * 30.44)
                
                result_entry = {
                    'miner_number': display_number,
                    'model': model,
                    'quantity': quantity,
                    'power_consumption': power_consumption,
                    'electricity_cost': electricity_cost,
                    'machine_price': machine_price,
                    'total_machine_cost': total_investment,
                    'daily_profit': calc_result.get('daily_profit', 0),
                    'daily_revenue': calc_result.get('daily_revenue', 0),
                    'daily_cost': calc_result.get('daily_cost', 0),
                    'monthly_profit': monthly_profit,
                    'roi_days': roi_days,
                    'hash_rate': calc_result.get('hashrate', hashrate),
                    'daily_btc': calc_result.get('daily_btc', 0)
                }
                
                results.append(result_entry)
                
                # 累加统计
                total_daily_profit += result_entry['daily_profit']
                total_daily_revenue += result_entry['daily_revenue']
                total_daily_cost += result_entry['daily_cost']
                
            except Exception as e:
                logger.error(f"计算矿机组 {model} 失败: {e}")
                continue
        
        # 创建汇总信息
        summary = {
            'total_miners': total_miners,
            'total_daily_profit': round(total_daily_profit, 2),
            'total_daily_revenue': round(total_daily_revenue, 2),
            'total_daily_cost': round(total_daily_cost, 2),
            'total_monthly_profit': round(total_daily_profit * 30, 2),
            'unique_groups': len(results),
            'average_roi_days': round(sum(r.get('roi_days', 0) for r in results) / len(results), 1) if results else 0
        }
        
        logger.info(f"批量计算完成: {len(results)} 个组，共 {total_miners} 台矿机")
        
        return jsonify({
            'success': True,
            'results': results,
            'summary': summary,
            'optimization_info': {
                'original_entries': len(miners),
                'optimized_groups': len(results),
                'total_miners': total_miners
            }
        })
        
    except Exception as e:
        logger.error(f"批量计算失败: {e}")
        return jsonify({
            'success': False,
            'error': 'calculation_failed',
            'message': '批量计算失败，请检查输入参数'
        }), 500

def get_shared_network_data():
    """获取共享的网络数据"""
    try:
        latest_data = NetworkSnapshot.query.order_by(
            NetworkSnapshot.recorded_at.desc()
        ).first()
        
        if latest_data:
            return {
                'btc_price': latest_data.btc_price,
                'network_difficulty': latest_data.network_difficulty,
                'network_hashrate': latest_data.network_hashrate,
                'block_reward': latest_data.block_reward
            }
        else:
            # 默认值
            return {
                'btc_price': 43000.0,
                'network_difficulty': 83148355189239.0,
                'network_hashrate': 650.0,
                'block_reward': 3.125
            }
            
    except Exception as e:
        logger.warning(f"获取网络数据失败: {e}")
        return {
            'btc_price': 43000.0,
            'network_difficulty': 83148355189239.0,
            'network_hashrate': 650.0,
            'block_reward': 3.125
        }

@batch_bp.route('/api/export/csv', methods=['POST'])
def export_csv():
    """导出批量计算结果为CSV"""
    try:
        data = request.get_json()
        results = data.get('results', [])
        
        if not results:
            return jsonify({
                'success': False,
                'message': '没有数据可导出'
            }), 400
        
        # 生成CSV内容
        csv_lines = [
            'Miner Number,Model,Quantity,Daily Revenue ($),Daily Cost ($),Daily Profit ($),'
            'Monthly Profit ($),ROI Days,Hash Rate (TH/s),Power (W),Machine Price ($),'
            'Total Cost ($),Daily BTC,Electricity Cost ($/kWh)'
        ]
        
        for result in results:
            line = ','.join([
                f'"{result.get("miner_number", "")}"',
                f'"{result.get("model", "")}"',
                str(result.get('quantity', 0)),
                f"{result.get('daily_revenue', 0):.2f}",
                f"{result.get('daily_cost', 0):.2f}",
                f"{result.get('daily_profit', 0):.2f}",
                f"{result.get('monthly_profit', 0):.2f}",
                str(result.get('roi_days', 0)),
                f"{result.get('hash_rate', 0):.1f}",
                f"{result.get('power_consumption', 0):.0f}",
                f"{result.get('machine_price', 0):.2f}",
                f"{result.get('total_machine_cost', 0):.2f}",
                f"{result.get('daily_btc', 0):.8f}",
                f"{result.get('electricity_cost', 0):.3f}"
            ])
            csv_lines.append(line)
        
        csv_content = '\n'.join(csv_lines)
        
        # 返回CSV文件
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename="mining_batch_results_{len(results)}_miners_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
        return response
        
    except Exception as e:
        logger.error(f"CSV导出失败: {e}")
        return jsonify({
            'success': False,
            'message': 'CSV导出失败'
        }), 500

@batch_bp.route('/api/export/excel', methods=['POST'])
def export_excel():
    """导出批量计算结果为Excel"""
    if not EXCEL_AVAILABLE:
        return jsonify({
            'success': False,
            'message': 'Excel导出功能不可用'
        }), 503
    
    try:
        data = request.get_json()
        results = data.get('results', [])
        summary = data.get('summary', {})
        
        if not results:
            return jsonify({
                'success': False,
                'message': '没有数据可导出'
            }), 400
        
        # 创建Excel工作簿
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Mining Batch Results"
        
        # 设置标题样式
        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")
        
        # 添加标题
        worksheet['A1'] = f"Bitcoin Mining Batch Calculation Results - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        worksheet['A1'].font = title_font
        worksheet.merge_cells('A1:M1')
        
        # 添加汇总信息
        summary_row = 3
        worksheet[f'A{summary_row}'] = "Summary:"
        worksheet[f'A{summary_row}'].font = header_font
        
        summary_data = [
            ('Total Miners:', summary.get('total_miners', 0)),
            ('Total Daily Profit:', f"${summary.get('total_daily_profit', 0):.2f}"),
            ('Total Monthly Profit:', f"${summary.get('total_monthly_profit', 0):.2f}"),
            ('Average ROI (days):', summary.get('average_roi_days', 0))
        ]
        
        for i, (label, value) in enumerate(summary_data):
            worksheet[f'A{summary_row + i + 1}'] = label
            worksheet[f'B{summary_row + i + 1}'] = value
        
        # 添加表头
        headers_row = summary_row + len(summary_data) + 2
        headers = [
            'Miner Number', 'Model', 'Quantity', 'Daily Revenue ($)', 'Daily Cost ($)',
            'Daily Profit ($)', 'Monthly Profit ($)', 'ROI Days', 'Hash Rate (TH/s)',
            'Power (W)', 'Machine Price ($)', 'Total Cost ($)', 'Daily BTC'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=headers_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 添加数据
        for row_idx, result in enumerate(results, headers_row + 1):
            worksheet.cell(row=row_idx, column=1, value=result.get('miner_number', ''))
            worksheet.cell(row=row_idx, column=2, value=result.get('model', ''))
            worksheet.cell(row=row_idx, column=3, value=result.get('quantity', 0))
            worksheet.cell(row=row_idx, column=4, value=round(result.get('daily_revenue', 0), 2))
            worksheet.cell(row=row_idx, column=5, value=round(result.get('daily_cost', 0), 2))
            worksheet.cell(row=row_idx, column=6, value=round(result.get('daily_profit', 0), 2))
            worksheet.cell(row=row_idx, column=7, value=round(result.get('monthly_profit', 0), 2))
            worksheet.cell(row=row_idx, column=8, value=result.get('roi_days', 0))
            worksheet.cell(row=row_idx, column=9, value=round(result.get('hash_rate', 0), 1))
            worksheet.cell(row=row_idx, column=10, value=round(result.get('power_consumption', 0), 0))
            worksheet.cell(row=row_idx, column=11, value=round(result.get('machine_price', 0), 2))
            worksheet.cell(row=row_idx, column=12, value=round(result.get('total_machine_cost', 0), 2))
            worksheet.cell(row=row_idx, column=13, value=round(result.get('daily_btc', 0), 8))
        
        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # 保存到内存
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        # 返回Excel文件
        filename = f"mining_batch_results_{len(results)}_miners_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Excel导出失败: {e}")
        return jsonify({
            'success': False,
            'message': 'Excel导出失败'
        }), 500

@batch_bp.route('/api/template/download')
def download_template():
    """下载批量计算模板"""
    try:
        # 获取活跃矿机列表作为示例
        miners = MinerModel.get_active_miners()
        
        # 创建CSV模板
        csv_lines = [
            'miner_number,model,quantity,electricity_cost,hashrate,power_consumption,machine_price,decay_rate'
        ]
        
        # 添加示例数据
        for i, miner in enumerate(miners[:5], 1):  # 最多5个示例
            line = ','.join([
                f'Miner-{i:03d}',
                miner.model_name,
                '1',
                '0.06',
                str(miner.reference_hashrate),
                str(miner.reference_power),
                str(miner.reference_price or 2500),
                '0.5'
            ])
            csv_lines.append(line)
        
        csv_content = '\n'.join(csv_lines)
        
        # 返回模板文件
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = 'attachment; filename="batch_calculator_template.csv"'
        return response
        
    except Exception as e:
        logger.error(f"模板下载失败: {e}")
        return jsonify({
            'success': False,
            'message': '模板下载失败'
        }), 500

@batch_bp.route('/api/upload/csv', methods=['POST'])
def upload_csv():
    """上传CSV文件进行批量计算"""
    try:
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'message': '没有上传文件'
            }), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': '没有选择文件'
            }), 400
        
        if not file.filename.lower().endswith('.csv'):
            return jsonify({
                'success': False,
                'message': '只支持CSV文件'
            }), 400
        
        # 读取CSV内容
        content = file.read().decode('utf-8')
        lines = content.strip().split('\n')
        
        if len(lines) < 2:
            return jsonify({
                'success': False,
                'message': 'CSV文件格式无效'
            }), 400
        
        # 解析CSV头部
        headers = [h.strip().lower() for h in lines[0].split(',')]
        required_headers = ['model', 'quantity', 'electricity_cost']
        
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            return jsonify({
                'success': False,
                'message': f'缺少必需的列: {", ".join(missing_headers)}'
            }), 400
        
        # 解析数据行
        miners = []
        for line_num, line in enumerate(lines[1:], 2):
            try:
                values = [v.strip() for v in line.split(',')]
                if len(values) != len(headers):
                    continue
                
                miner_data = dict(zip(headers, values))
                
                # 验证和转换数据
                miner = {
                    'model': miner_data.get('model', 'Antminer S19 Pro'),
                    'quantity': int(float(miner_data.get('quantity', 1))),
                    'electricity_cost': float(miner_data.get('electricity_cost', 0.06)),
                    'miner_number': miner_data.get('miner_number', ''),
                    'hashrate': float(miner_data.get('hashrate', 0)) if miner_data.get('hashrate') else 0,
                    'power_consumption': float(miner_data.get('power_consumption', 0)) if miner_data.get('power_consumption') else 0,
                    'machine_price': float(miner_data.get('machine_price', 2500)) if miner_data.get('machine_price') else 2500,
                    'decay_rate': float(miner_data.get('decay_rate', 0)) if miner_data.get('decay_rate') else 0
                }
                
                miners.append(miner)
                
            except (ValueError, IndexError) as e:
                logger.warning(f"跳过无效数据行 {line_num}: {e}")
                continue
        
        if not miners:
            return jsonify({
                'success': False,
                'message': 'CSV文件中没有有效数据'
            }), 400
        
        return jsonify({
            'success': True,
            'miners': miners,
            'count': len(miners),
            'message': f'成功解析 {len(miners)} 条记录'
        })
        
    except Exception as e:
        logger.error(f"CSV上传失败: {e}")
        return jsonify({
            'success': False,
            'message': 'CSV文件上传失败'
        }), 500